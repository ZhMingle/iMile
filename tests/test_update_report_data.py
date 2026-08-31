from pathlib import Path
from collections import Counter
import tempfile
import unittest

import pandas as pd
from openpyxl import Workbook

import update_report_data


class UpdateReportDataTests(unittest.TestCase):
    def test_blank_display_alias_does_not_count_blank_station_rows(self):
        counts = Counter({"": 3448, "AKL": 4790})

        self.assertEqual(
            update_report_data.station_count_with_display_alias(
                counts,
                "PMN",
                "",
            ),
            0,
        )

    def test_real_display_alias_is_counted_once(self):
        counts = Counter({"HMT": 10, "Hamilton": 2})

        self.assertEqual(
            update_report_data.station_count_with_display_alias(
                counts,
                "HMT",
                "Hamilton",
            ),
            12,
        )

    def test_new_plymouth_v2_is_distinct_from_palmerston_north_v2(self):
        counts = Counter({"NPMV2": 205, "PMNV2": 770})

        self.assertEqual(update_report_data.station_count(counts, "NPMV2"), 205)
        self.assertEqual(update_report_data.station_count(counts, "PMN"), 770)

    def test_auckland_route_counts_exclude_other_and_blank_stations(self):
        frame = pd.DataFrame(
            {
                "路由码": ["301", "301", ""],
                "派件网点简码": ["AKL", "HMT", ""],
            }
        )

        self.assertEqual(
            update_report_data.auckland_route_counts(frame),
            Counter({"301": 1}),
        )

    def test_total_breakdown_keeps_unassigned_orders_once(self):
        workbook = Workbook()
        worksheet = workbook.active

        update_report_data.write_total_breakdown(
            worksheet,
            auckland_total=4790,
            non_auckland_total=0,
            source_total=8238,
        )

        self.assertEqual(
            worksheet.cell(13, 6).value,
            "当天总量（奥克兰 + 外省 + 未分配）",
        )
        self.assertEqual(worksheet.cell(14, 6).value, "8238(4790+0+3448)")

    def test_clean_route_code_normalizes_excel_integer_decimals(self):
        self.assertEqual(update_report_data.clean_route_code("306.0"), "306")
        self.assertEqual(update_report_data.clean_route_code(" 403.00 "), "403")

    def test_clean_route_code_preserves_meaningful_decimal_or_text_codes(self):
        self.assertEqual(update_report_data.clean_route_code("201.5"), "201.5")
        self.assertEqual(update_report_data.clean_route_code("201A"), "201A")
        self.assertEqual(update_report_data.clean_route_code("501 C"), "501C")

    def test_404a_supplier_is_overridden_to_feng(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "奥克兰"
        worksheet.cell(3, 1).value = "404 A"
        worksheet.cell(3, 7).value = "PANDA"
        worksheet.cell(4, 1).value = "总计"

        update_report_data.update_auckland_sheet(
            workbook,
            Counter({"404A": 10}),
        )

        self.assertEqual(worksheet.cell(3, 7).value, "Feng")

    def test_501c_supplier_is_overridden_to_panda(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "奥克兰"
        worksheet.cell(3, 1).value = "501 C"
        worksheet.cell(3, 7).value = "EMPIRE COURIER"
        worksheet.cell(4, 1).value = "总计"

        update_report_data.update_auckland_sheet(
            workbook,
            Counter({"501C": 4}),
        )

        self.assertEqual(worksheet.cell(3, 3).value, 4)
        self.assertEqual(worksheet.cell(3, 7).value, "PANDA")

    def test_3l_board_forecast_uses_200_piece_capacity(self):
        workbook = Workbook()
        worksheet = workbook.active
        arrivals = {
            "HMT": 1570,
            "TRG": 1190,
            "WLTV2": 1574,
            "NPL": 494,
            "PMN": 679,
            "TPO": 201,
            "RTR": 440,
            "WGR": 417,
            "HST": 345,
            "NPMV2": 195,
            "WGU": 166,
            "GSB": 88,
        }
        for row, (station, arrival) in enumerate(arrivals.items(), start=3):
            worksheet.cell(row, 1).value = station
            worksheet.cell(row, 3).value = arrival
        worksheet.cell(15, 1).value = "总计"
        worksheet.cell(2, 8).value = 135
        worksheet.cell(2, 9).value = "3L预测板数"
        worksheet.cell(13, 8).value = 350
        worksheet.cell(13, 9).value = "5L预测板数"
        for row in range(3, 12):
            worksheet.row_dimensions[row].height = 18
        worksheet.row_dimensions[13].height = 22.5
        worksheet.row_dimensions[14].height = 22.5
        worksheet.row_dimensions[15].height = 31

        update_report_data.write_board_forecast_values(worksheet)
        update_report_data.write_board_forecast_values(worksheet)

        self.assertEqual(worksheet.column_dimensions["H"].width, 16)
        self.assertEqual(worksheet.column_dimensions["I"].width, 24)
        self.assertEqual(worksheet.cell(1, 8).value, 200)
        self.assertEqual(
            [worksheet.cell(row, 8).value for row in range(2, 11)],
            ["HMT", "TRG/RTR", "TPO", "NPL/HST", "PMN", "WLTV2", "NPMV2", "WGU", "GSB"],
        )
        self.assertEqual(
            [worksheet.cell(row, 9).value for row in range(2, 11)],
            [7.85, "5.95/2.2(8.15)", 1.0, "2.47/1.73(4.2)", 3.4, 7.87, 0.97, 0.83, 0.44],
        )
        self.assertEqual(worksheet.cell(11, 8).value, "总计")
        self.assertEqual(worksheet.cell(11, 9).value, 34.71)
        self.assertEqual(
            [worksheet.row_dimensions[row].height for row in range(2, 12)],
            [18] * 10,
        )
        self.assertIsNone(worksheet.cell(12, 8).value)
        self.assertEqual(worksheet.cell(13, 8).value, 350)
        self.assertEqual(worksheet.cell(13, 9).value, "5L预测板数")
        self.assertEqual(
            [worksheet.cell(row, 8).value for row in range(14, 23)],
            ["HMT", "TRG/RTR", "TPO", "NPL/HST", "PMN", "WLTV2", "NPMV2", "WGU", "GSB"],
        )
        self.assertEqual(
            [worksheet.cell(row, 9).value for row in range(14, 23)],
            [4.49, "3.4/1.26(4.66)", 0.57, "1.41/0.99(2.4)", 1.94, 4.5, 0.56, 0.47, 0.25],
        )
        self.assertEqual(worksheet.cell(23, 8).value, "总计")
        self.assertEqual(worksheet.cell(23, 9).value, 19.84)
        self.assertEqual(
            [worksheet.row_dimensions[row].height for row in range(13, 24)],
            [22.5] * 11,
        )
        self.assertEqual(
            update_report_data.BOARD_FORECAST_GROUPS,
            [
                ("HMT",),
                ("TRG", "RTR"),
                ("TPO",),
                ("NPL", "HST"),
                ("PMN",),
                ("WLTV2",),
                ("NPMV2",),
                ("WGU",),
                ("GSB",),
            ],
        )

    def test_non_auckland_arrivals_follow_dispatch_distance_order(self):
        self.assertEqual(
            update_report_data.NON_AUCKLAND_STATIONS,
            [
                "HMT",
                "TRG",
                "RTR",
                "TPO",
                "NPL",
                "HST",
                "PMN",
                "WLTV2",
                "WGR",
                "NPMV2",
                "WGU",
                "GSB",
            ],
        )

    def test_new_stations_move_summary_below_the_expanded_overview(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "非奥克兰"
        old_stations = ["HMT", "TRG", "WLTV2", "NPL", "PMN", "RTR", "WGR", "HST"]
        for row, station in enumerate(old_stations, start=3):
            worksheet.cell(row, 1).value = station
        worksheet.cell(11, 1).value = "总计"
        worksheet.cell(13, 1).value = "Aliexpress 单量"
        worksheet.cell(13, 3).value = "顺友单量"
        worksheet.cell(13, 6).value = "当天总量（奥克兰 + 外省 + 未分配）"
        worksheet.cell(14, 1).value = 0
        worksheet.cell(14, 3).value = 0
        worksheet.cell(14, 6).value = "0(0+0)"
        for col in (1, 3, 6):
            worksheet.cell(13, col).number_format = "@"
            worksheet.cell(14, col).number_format = "0"
        worksheet.cell(13, 8).value = 350

        station_counts = Counter(
            {
                "HMT": 2045,
                "TRG": 1529,
                "RTR": 602,
                "TPO": 288,
                "NPL": 608,
                "HST": 471,
                "PMNV2": 770,
                "WLTV2": 1414,
                "WGR": 543,
                "NPMV2": 205,
                "WGU": 174,
                "GSB": 123,
            }
        )

        non_auckland_total = update_report_data.update_non_auckland_sheet(
            workbook,
            station_counts=station_counts,
            cainiao_counts=Counter(),
            sunyou_counts=Counter(),
            aliexpress_count=0,
            sunyou_count=0,
            auckland_total=7268,
            source_total=16042,
        )

        self.assertEqual(non_auckland_total, 8772)
        self.assertEqual(
            [worksheet.cell(row, 1).value for row in range(3, 15)],
            update_report_data.NON_AUCKLAND_STATIONS,
        )
        self.assertEqual(worksheet.cell(15, 1).value, "总计")
        self.assertEqual(worksheet.cell(15, 3).value, 8772)
        self.assertEqual(worksheet.cell(16, 1).value, "Aliexpress 单量")
        self.assertEqual(worksheet.cell(16, 3).value, "顺友单量")
        self.assertEqual(worksheet.cell(16, 1).number_format, "@")
        self.assertEqual(worksheet.cell(17, 1).number_format, "0")
        self.assertEqual(
            worksheet.cell(16, 6).value,
            "当天总量（奥克兰 + 外省 + 未分配）",
        )
        self.assertEqual(worksheet.cell(17, 6).value, "16042(7268+8772+2)")
        self.assertEqual(worksheet.cell(1, 8).value, 200)
        self.assertEqual(
            [worksheet.cell(row, 8).value for row in range(2, 11)],
            ["HMT", "TRG/RTR", "TPO", "NPL/HST", "PMN", "WLTV2", "NPMV2", "WGU", "GSB"],
        )
        self.assertEqual(worksheet.cell(11, 8).value, "总计")
        self.assertIsNone(worksheet.cell(12, 8).value)
        self.assertEqual(worksheet.cell(13, 8).value, 350)
        self.assertEqual(worksheet.cell(13, 9).value, "5L预测板数")
        self.assertEqual(
            [worksheet.cell(row, 8).value for row in range(14, 23)],
            ["HMT", "TRG/RTR", "TPO", "NPL/HST", "PMN", "WLTV2", "NPMV2", "WGU", "GSB"],
        )
        self.assertEqual(worksheet.cell(23, 8).value, "总计")

    def test_expanded_summary_layout_is_idempotent_when_merged(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "非奥克兰"
        for row, station in enumerate(update_report_data.NON_AUCKLAND_STATIONS, start=3):
            worksheet.cell(row, 1).value = station
        worksheet.cell(15, 1).value = "总计"
        worksheet.cell(16, 1).value = "Aliexpress 单量"
        worksheet.cell(16, 6).value = "当天总量（奥克兰 + 外省 + 未分配）"
        worksheet.cell(17, 6).value = "16042(7268+8772+2)"
        worksheet.merge_cells("F16:G16")
        worksheet.merge_cells("F17:G17")

        for _ in range(2):
            self.assertEqual(
                update_report_data.ensure_non_auckland_station_rows(worksheet),
                15,
            )

        self.assertEqual(worksheet.cell(16, 1).value, "Aliexpress 单量")
        self.assertEqual(worksheet.cell(17, 6).value, "16042(7268+8772+2)")
        self.assertEqual(
            {str(merged_range) for merged_range in worksheet.merged_cells.ranges},
            {"F16:G16", "F17:G17"},
        )

    def test_manual_english_export_headers_are_canonicalized(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "Central Waybill Query.xlsx"
            pd.DataFrame(
                {
                    "Waybill Number": ["6072126000001"],
                    "Routing Code": ["305"],
                    "Delivery Station S-Code": ["AKL"],
                    "Client Code": ["C2103960401"],
                }
            ).to_excel(source, index=False)

            frame = update_report_data.read_source_xlsx(source)

            self.assertEqual(list(frame.columns), update_report_data.REQUIRED_COLUMNS)
            self.assertEqual(frame.loc[0, "运单号"], "6072126000001")
            self.assertEqual(frame.loc[0, "派件网点简码"], "AKL")


if __name__ == "__main__":
    unittest.main()
