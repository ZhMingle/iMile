import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path
import warnings
import re
import zipfile
import xml.etree.ElementTree as ET
from copy import copy

import pandas as pd
from openpyxl import load_workbook

from report_config import BOARD_3L_CAPACITY, BOARD_5L_CAPACITY
from report_source_freshness import center_waybill_file_freshness_warning


SOURCE_PATTERN = "*中心运单查询*.xlsx"
TEMPLATE_PATTERN = "当日数据统计*.xlsx"
OUTPUT_FILE = Path("当日数据统计.xlsx")
today = datetime.now()
REPORT_DATE = f"{today.month}月{today.day:02d}日"

SOURCE_SHEET = "数据源1-预测"
BASE_COLUMNS = ["运单号", "路由码", "派件网点简码"]
MERCHANT_COLUMN = "商家编号"
REQUIRED_COLUMNS = [*BASE_COLUMNS, MERCHANT_COLUMN]
TEMPLATE_REQUIRED_COLUMNS = BASE_COLUMNS
SOURCE_COLUMN_KEY_MAP = {
    "运单号": "运单号",
    "waybillnumber": "运单号",
    "trackingnumber": "运单号",
    "路由码": "路由码",
    "routingcode": "路由码",
    "routecode": "路由码",
    "派件网点简码": "派件网点简码",
    "deliverystationscode": "派件网点简码",
    "deliverystationcode": "派件网点简码",
    "dispatchstationcode": "派件网点简码",
    "商家编号": "商家编号",
    "clientcode": "商家编号",
    "merchantcode": "商家编号",
}

MERCHANT_CODES = {
    "TEMU": "C2103951401",
    "CAINIAO": "C2103960401",
    "SUNYOU": "C2104258001",
}
CAINIAO_MERCHANT_CODE = MERCHANT_CODES["CAINIAO"]
SUNYOU_MERCHANT_CODE = MERCHANT_CODES["SUNYOU"]

NON_AUCKLAND_STATIONS = ["HMT", "TRG", "RTR", "TPO", "NPL", "HST", "PMN", "WLTV2", "WGR"]
BOARD_FORECAST_STATIONS = ["HMT", "TRG", "RTR", "TPO", "NPL", "HST", "PMN", "WLTV2"]
STATION_ALIASES = {
    "WLTV2": ["WLTV2", "WLT", "AKL-DC"],
    "PMN": ["PMN", "PMNV2", "Palmerston NorthV2"],
}
STATION_DISPLAY_ALIASES = {
    "HMT": "Hamilton",
    "TRG": "Tauranga",
    "NPL": "Napier",
    "RTR": " Rotorua",
}
AUCKLAND_ROUTE_SUPPLIERS = {
    "404A": "Feng",
    "501C": "PANDA",
    "406": "Feng",
    "601": "Good Day Removals Ltd",
}


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_route_code(value):
    route_code = re.sub(r"\s+", "", clean_text(value))
    # Excel exports sometimes store numeric route codes as decimals (for
    # example, 306 becomes 306.0).  The report template stores the same code as
    # text, so normalize only an all-zero decimal suffix before matching.
    return re.sub(r"^(\d+)\.0+$", r"\1", route_code)


def canonical_source_column(value):
    original = clean_text(value)
    key = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", original.lower())
    return SOURCE_COLUMN_KEY_MAP.get(key, original)


def is_route_code(value):
    return value == "HST" or any(char.isdigit() for char in value)


def load_source_data(source_file=None, allow_old_source=False):
    if source_file is None:
        source_file, df = find_latest_source_file()
    else:
        source_file = Path(source_file)
        if not source_file.exists():
            raise FileNotFoundError(f"Source file not found: {source_file}")
        df = read_source_xlsx(source_file)
        missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
        if missing:
            raise ValueError(f"Source file is missing required columns: {source_file}: {missing}")

    freshness_warning = center_waybill_file_freshness_warning(source_file)
    if freshness_warning and not allow_old_source:
        raise RuntimeError(
            f"{freshness_warning}\n\n"
            "为防止误发，操作已停止。请更新文件后重试；"
            "如果确实要处理旧文件，请显式使用 --allow-old-source。"
        )

    print(f"Using source file: {source_file}")
    df["路由码"] = df["路由码"].map(clean_route_code)
    df = df[df["运单号"] != ""].drop_duplicates(subset=["运单号"], keep="first")
    return df.reset_index(drop=True)


def find_latest_source_file():
    files = [
        path
        for path in Path(".").glob(SOURCE_PATTERN)
        if not path.name.startswith("~$")
    ]
    if not files:
        raise FileNotFoundError(f"No source file found matching {SOURCE_PATTERN}")

    skipped = []
    for path in sorted(files, key=lambda item: item.stat().st_mtime, reverse=True):
        df = read_source_xlsx(path)
        missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
        if not missing:
            if skipped:
                print("Skipped source files missing required columns:")
                for skipped_path, skipped_missing in skipped:
                    print(f"- {skipped_path}: missing {skipped_missing}")
            return path, df
        skipped.append((path, missing))

    details = "; ".join(f"{path}: missing {missing}" for path, missing in skipped)
    raise ValueError(f"No valid source file found matching {SOURCE_PATTERN}. {details}")


def find_latest_template_file():
    canonical_template = Path("当日数据统计_20260603_233359_公式版.xlsx")
    if canonical_template.exists():
        return canonical_template

    files = [
        path
        for path in Path(".").glob(TEMPLATE_PATTERN)
        if not path.name.startswith("~$") and not re.search(r"_\d{8}_\d{6}\.xlsx$", path.name)
    ]
    if not files:
        raise FileNotFoundError(f"No template file found matching {TEMPLATE_PATTERN}")
    return max(files, key=lambda path: path.stat().st_mtime)


def read_source_xlsx(path):
    shared_strings = []
    with zipfile.ZipFile(path) as archive:
        if "xl/sharedStrings.xml" in archive.namelist():
            with archive.open("xl/sharedStrings.xml") as shared_stream:
                for event, element in ET.iterparse(shared_stream, events=("end",)):
                    if element.tag.endswith("}si"):
                        shared_strings.append("".join(text.text or "" for text in element.iter() if text.tag.endswith("}t")))
                        element.clear()

        sheet_name = get_first_sheet_path(archive)
        required_indexes = None
        rows = []

        with archive.open(sheet_name) as sheet_stream:
            for event, row_element in ET.iterparse(sheet_stream, events=("end",)):
                if not row_element.tag.endswith("}row"):
                    continue

                row_number = int(row_element.attrib.get("r", "0"))
                values_by_col = {}
                for cell in row_element:
                    if not cell.tag.endswith("}c"):
                        continue
                    col_index = column_index_from_cell_ref(cell.attrib.get("r", ""))
                    values_by_col[col_index] = read_cell_value(cell, shared_strings)

                if row_number == 1:
                    header_to_index = {
                        canonical_source_column(value): col for col, value in values_by_col.items()
                    }
                    required_indexes = {column: header_to_index[column] for column in REQUIRED_COLUMNS if column in header_to_index}
                elif required_indexes:
                    row = {
                        column: clean_text(values_by_col.get(col_index, ""))
                        for column, col_index in required_indexes.items()
                    }
                    if row.get("运单号"):
                        rows.append(row)

                row_element.clear()

    return pd.DataFrame(rows, columns=[column for column in REQUIRED_COLUMNS if rows and column in rows[0]] or REQUIRED_COLUMNS)


def get_first_sheet_path(archive):
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationship_by_id = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    first_sheet = next(element for element in workbook.iter() if element.tag.endswith("}sheet"))
    relationship_id = first_sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
    target = relationship_by_id[relationship_id]
    if target.startswith("/"):
        return target.lstrip("/")
    return f"xl/{target.lstrip('/')}"


def column_index_from_cell_ref(cell_ref):
    letters = re.match(r"[A-Z]+", cell_ref).group(0)
    number = 0
    for letter in letters:
        number = number * 26 + ord(letter) - ord("A") + 1
    return number


def read_cell_value(cell, shared_strings):
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.iter() if text.tag.endswith("}t"))

    value_element = next((child for child in cell if child.tag.endswith("}v")), None)
    if value_element is None or value_element.text is None:
        return ""

    if cell_type == "s":
        index = int(value_element.text)
        return shared_strings[index] if index < len(shared_strings) else ""

    return value_element.text


def clear_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def ensure_data_source_column_order(ws):
    desired_first_columns = TEMPLATE_REQUIRED_COLUMNS
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if headers[: len(desired_first_columns)] == desired_first_columns:
        return

    header_to_col = {header: col for col, header in enumerate(headers, start=1) if header}
    missing = [column for column in desired_first_columns if column not in header_to_col]
    if missing:
        raise ValueError(f"{SOURCE_SHEET} is missing columns: {missing}")

    ordered_columns = [header_to_col[column] for column in desired_first_columns]
    ordered_columns.extend(
        col
        for col, header in enumerate(headers, start=1)
        if header not in desired_first_columns
    )

    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for row_index in range(1, max_row + 1):
        row_snapshot = []
        for source_col in ordered_columns:
            source_cell = ws.cell(row_index, source_col)
            row_snapshot.append(
                {
                    "value": source_cell.value,
                    "fill": copy(source_cell.fill),
                    "font": copy(source_cell.font),
                    "border": copy(source_cell.border),
                    "alignment": copy(source_cell.alignment),
                    "number_format": source_cell.number_format,
                    "protection": copy(source_cell.protection),
                }
            )
        rows.append(row_snapshot)

    column_widths = {
        target_col: ws.column_dimensions[ws.cell(1, source_col).column_letter].width
        for target_col, source_col in enumerate(ordered_columns, start=1)
    }

    clear_range(ws, 1, max_row, 1, max_col)
    for row_index, row_snapshot in enumerate(rows, start=1):
        for col_index, cell_snapshot in enumerate(row_snapshot, start=1):
            target_cell = ws.cell(row_index, col_index)
            target_cell.value = cell_snapshot["value"]
            target_cell.fill = copy(cell_snapshot["fill"])
            target_cell.font = copy(cell_snapshot["font"])
            target_cell.border = copy(cell_snapshot["border"])
            target_cell.alignment = copy(cell_snapshot["alignment"])
            target_cell.number_format = cell_snapshot["number_format"]
            target_cell.protection = copy(cell_snapshot["protection"])

    for col_index, width in column_widths.items():
        ws.column_dimensions[ws.cell(1, col_index).column_letter].width = width


def write_data_source(wb, df):
    ws = wb[SOURCE_SHEET]
    ensure_data_source_column_order(ws)

    header_to_col = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    missing = [column for column in TEMPLATE_REQUIRED_COLUMNS if column not in header_to_col]
    if missing:
        raise ValueError(f"{SOURCE_SHEET} is missing columns: {missing}")

    old_max_row = ws.max_row
    if old_max_row > 1:
        clear_range(ws, 2, old_max_row, 1, ws.max_column)

    for index, row in df.iterrows():
        excel_row = index + 2
        ws.cell(excel_row, header_to_col["运单号"]).value = row["运单号"]
        ws.cell(excel_row, header_to_col["派件网点简码"]).value = row["派件网点简码"]
        ws.cell(excel_row, header_to_col["路由码"]).value = row["路由码"]
        if MERCHANT_COLUMN in header_to_col:
            ws.cell(excel_row, header_to_col[MERCHANT_COLUMN]).value = row[MERCHANT_COLUMN]

    first_extra_row = len(df) + 2
    if old_max_row >= first_extra_row:
        ws.delete_rows(first_extra_row, old_max_row - first_extra_row + 1)


def find_total_row(ws, column=1, label="总计"):
    for row in range(1, ws.max_row + 1):
        value = clean_text(ws.cell(row, column).value)
        if value == label:
            return row
    raise ValueError(f"Could not find {label} in {ws.title}")


def copy_cell_style(source_cell, target_cell):
    for attr in ["fill", "font", "border", "alignment", "number_format", "protection"]:
        setattr(target_cell, attr, copy(getattr(source_cell, attr)))


def station_count(counts, station):
    aliases = STATION_ALIASES.get(station, [station])
    return sum(counts.get(alias, 0) for alias in aliases)


def station_count_with_display_alias(counts, station, display_alias):
    keys = {
        clean_text(value)
        for value in [*STATION_ALIASES.get(station, [station]), display_alias]
        if clean_text(value)
    }
    return sum(counts.get(key, 0) for key in keys)


def auckland_route_counts(df):
    return Counter(
        df.loc[df["派件网点简码"].eq("AKL"), "路由码"]
    )


def ensure_non_auckland_station_rows(ws):
    current_total_row = find_total_row(ws, column=1)
    target_total_row = 3 + len(NON_AUCKLAND_STATIONS)
    total_label = ws.cell(current_total_row, 1).value
    existing_station_values = {}
    for row in range(3, current_total_row):
        station = clean_text(ws.cell(row, 1).value)
        if station == "WLT":
            station = "WLTV2"
        if station:
            existing_station_values[station] = [ws.cell(row, col).value for col in range(4, 6)]
    total_styles = []
    for col in range(1, 8):
        source_cell = ws.cell(current_total_row, col)
        total_styles.append(
            {
                attr: copy(getattr(source_cell, attr))
                for attr in ["fill", "font", "border", "alignment", "number_format", "protection"]
            }
        )

    for row in range(3, target_total_row):
        style_row = row if row < current_total_row else max(3, current_total_row - 1)
        for col in range(1, 8):
            copy_cell_style(ws.cell(style_row, col), ws.cell(row, col))

    if current_total_row != target_total_row:
        for col in range(1, 8):
            ws.cell(target_total_row, col).value = None
            if current_total_row < target_total_row:
                ws.cell(current_total_row, col).value = None

    for col, style in enumerate(total_styles, start=1):
        target_cell = ws.cell(target_total_row, col)
        for attr, value in style.items():
            setattr(target_cell, attr, copy(value))

    for row, station in enumerate(NON_AUCKLAND_STATIONS, start=3):
        ws.cell(row, 1).value = station
        ws.cell(row, 2).value = STATION_DISPLAY_ALIASES.get(station)
        for col, value in zip(range(4, 6), existing_station_values.get(station, [None, None])):
            ws.cell(row, col).value = value

    ws.cell(target_total_row, 1).value = total_label
    return target_total_row


def rebuild_non_auckland_pivot(wb, station_counts):
    ws = wb["非奥克兰透视"]
    clear_range(ws, 1, max(ws.max_row, len(station_counts) + 4), 1, 2)

    ws.cell(3, 1).value = "派件网点简码"
    ws.cell(3, 2).value = "计数项:运单号"

    row = 4
    for station, count in sorted(station_counts.items()):
        ws.cell(row, 1).value = station
        ws.cell(row, 2).value = count
        row += 1

    ws.cell(row, 1).value = "总计"
    ws.cell(row, 2).value = sum(station_counts.values())


def rebuild_auckland_pivot(wb, route_counts):
    ws = wb["奥克兰透视"]
    clear_range(ws, 4, max(ws.max_row, len(route_counts) + 5), 1, 2)

    ws.cell(3, 1).value = "路由码"
    ws.cell(3, 2).value = "计数项:运单号"

    row = 4
    for route_code, count in sorted(route_counts.items(), key=lambda item: str(item[0])):
        ws.cell(row, 1).value = route_code
        ws.cell(row, 2).value = count
        row += 1

    ws.cell(row, 1).value = "总计"
    ws.cell(row, 2).value = sum(route_counts.values())

    # Refresh only the fixed side-summary tables. Scanning the whole sheet can
    # overwrite helper cells that happen to contain route-like values.
    route_table_ranges = [
        (8, 9, 6, 7),     # WGR
        (9, 24, 10, 11),  # HMT
        (30, 31, 7, 8),   # RTR
        (29, 34, 10, 11), # TRG
        (40, 47, 10, 11), # NPL/HST
        (9, 21, 13, 14),  # WLTV2
    ]
    for start_row, end_row, route_col, qty_col in route_table_ranges:
        for row in range(start_row, end_row + 1):
            route_code = clean_route_code(ws.cell(row, route_col).value)
            if is_route_code(route_code):
                ws.cell(row, qty_col).value = route_counts.get(route_code, 0)


def update_auckland_sheet(wb, route_counts):
    ws = wb["奥克兰"]
    ws.cell(1, 1).value = f"{REPORT_DATE}奥克兰分单"
    ws.cell(2, 12).value = REPORT_DATE
    total_row = find_total_row(ws, column=1)

    for row in range(3, total_row):
        route_code = clean_route_code(ws.cell(row, 1).value)
        if route_code:
            ws.cell(row, 3).value = route_counts.get(route_code, 0)
            if route_code in AUCKLAND_ROUTE_SUPPLIERS:
                ws.cell(row, 7).value = AUCKLAND_ROUTE_SUPPLIERS[route_code]

    total = sum(ws.cell(row, 3).value or 0 for row in range(3, total_row))
    ws.cell(total_row, 3).value = total
    return total


def update_non_auckland_sheet(
    wb,
    station_counts,
    cainiao_counts,
    sunyou_counts,
    aliexpress_count,
    sunyou_count,
    auckland_total,
    source_total,
):
    ws = wb["非奥克兰"]
    ws.cell(1, 1).value = f"{REPORT_DATE}非奥克兰到件货量"
    ws.cell(1, 9).value = REPORT_DATE
    total_row = ensure_non_auckland_station_rows(ws)

    for row in range(3, total_row):
        station = clean_text(ws.cell(row, 1).value)
        alias = clean_text(ws.cell(row, 2).value)

        arrival = station_count_with_display_alias(
            station_counts,
            station,
            alias,
        )
        if station == "RTR":
            arrival = sum(count for key, count in station_counts.items() if key.startswith("RTR"))

        ws.cell(row, 3).value = arrival
        ws.cell(row, 6).value = station_count(cainiao_counts, station)
        ws.cell(row, 7).value = station_count(sunyou_counts, station)

    non_auckland_total = sum(ws.cell(row, 3).value or 0 for row in range(3, total_row))
    cainiao_total = sum(ws.cell(row, 6).value or 0 for row in range(3, total_row))
    sunyou_total = sum(ws.cell(row, 7).value or 0 for row in range(3, total_row))

    ws.cell(total_row, 3).value = non_auckland_total
    ws.cell(total_row, 6).value = cainiao_total
    ws.cell(total_row, 7).value = sunyou_total
    ws.cell(14, 1).value = aliexpress_count
    ws.cell(13, 3).value = "顺友单量"
    ws.cell(14, 3).value = sunyou_count

    write_total_breakdown(
        ws,
        auckland_total,
        non_auckland_total,
        source_total,
    )

    write_board_forecast_values(ws)

    return non_auckland_total


def write_total_breakdown(
    ws,
    auckland_total,
    non_auckland_total,
    source_total=None,
):
    classified_total = auckland_total + non_auckland_total
    total = classified_total if source_total is None else source_total
    unassigned_total = total - classified_total
    if unassigned_total < 0:
        raise ValueError(
            "Auckland and non-Auckland totals exceed the unique source total; "
            "the report would double-count waybills."
        )

    source_header = ws.cell(13, 6)
    source_value = ws.cell(14, 6)
    header_style = {
        attr: copy(getattr(source_header, attr))
        for attr in ["fill", "font", "border", "alignment", "number_format", "protection"]
    }
    value_style = {
        attr: copy(getattr(source_value, attr))
        for attr in ["fill", "font", "border", "alignment", "number_format", "protection"]
    }

    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row <= 14
            and merged_range.max_row >= 13
            and merged_range.min_col <= 7
            and merged_range.max_col >= 4
        ):
            ws.unmerge_cells(str(merged_range))

    ws.cell(13, 4).value = None
    ws.cell(14, 4).value = None
    ws.cell(13, 5).value = None
    ws.cell(14, 5).value = None
    ws.cell(13, 7).value = None
    ws.cell(14, 7).value = None

    ws.merge_cells(start_row=13, start_column=6, end_row=13, end_column=7)
    ws.merge_cells(start_row=14, start_column=6, end_row=14, end_column=7)

    header_cell = ws.cell(13, 6)
    value_cell = ws.cell(14, 6)
    if unassigned_total:
        header_cell.value = "当天总量（奥克兰 + 外省 + 未分配）"
        value_cell.value = (
            f"{total}({auckland_total}+{non_auckland_total}"
            f"+{unassigned_total})"
        )
    else:
        header_cell.value = "当天总量（奥克兰 + 外省）"
        value_cell.value = f"{total}({auckland_total}+{non_auckland_total})"

    for attr, style in header_style.items():
        setattr(header_cell, attr, copy(style))
    for attr, style in value_style.items():
        setattr(value_cell, attr, copy(style))

    ws.column_dimensions["D"].width = 20.8181818181818
    ws.column_dimensions["E"].width = 12.8181818181818
    ws.column_dimensions["F"].width = 24.9
    ws.column_dimensions["G"].width = 24.9


def round_excel(value, digits=2):
    return round(float(value), digits)


def copy_board_row_style(ws, source_row, target_row):
    for col in (8, 9):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)
        for attr in ["fill", "font", "border", "alignment", "number_format", "protection"]:
            setattr(target_cell, attr, copy(getattr(source_cell, attr)))


def write_board_forecast_values(ws):
    # 3L forecasts now use 200 pieces per board.  Persist the assumption in
    # the report so both the workbook and the generated image show the same
    # capacity and use it for every station calculation.
    ws.cell(2, 8).value = BOARD_3L_CAPACITY
    base_3l = BOARD_3L_CAPACITY
    base_5l = ws.cell(13, 8).value or BOARD_5L_CAPACITY

    total_row = find_total_row(ws, column=1)
    arrival_by_station = {
        clean_text(ws.cell(row, 1).value): ws.cell(row, 3).value or 0
        for row in range(3, total_row)
    }

    rows_3l = {
        row: station
        for row, station in enumerate(BOARD_FORECAST_STATIONS, start=3)
    }
    rows_5l = {
        row: station
        for row, station in enumerate(BOARD_FORECAST_STATIONS, start=14)
    }

    copy_board_row_style(ws, 10, 11)
    copy_board_row_style(ws, 9, 10)
    copy_board_row_style(ws, 21, 22)
    copy_board_row_style(ws, 20, 21)

    for row, station in rows_3l.items():
        ws.cell(row, 8).value = station
        ws.cell(row, 9).value = round_excel(arrival_by_station.get(station, 0) / base_3l)

    ws.cell(11, 8).value = "总计"
    ws.cell(11, 9).value = sum(
        ws.cell(row, 9).value or 0
        for row in rows_3l
    )
    ws.cell(11, 9).value = round_excel(ws.cell(11, 9).value)

    for row, station in rows_5l.items():
        ws.cell(row, 8).value = station
        ws.cell(row, 9).value = round_excel(arrival_by_station.get(station, 0) / base_5l)

    ws.cell(22, 8).value = "总计"
    ws.cell(22, 9).value = sum(
        ws.cell(row, 9).value or 0
        for row in rows_5l
    )
    ws.cell(22, 9).value = round_excel(ws.cell(22, 9).value)


def main(source_file=None, allow_old_source=False):
    df = load_source_data(source_file, allow_old_source=allow_old_source)

    station_counts = Counter(df["派件网点简码"])
    route_counts = auckland_route_counts(df)
    cainiao_mask = df[MERCHANT_COLUMN].eq(CAINIAO_MERCHANT_CODE)
    sunyou_mask = df[MERCHANT_COLUMN].eq(SUNYOU_MERCHANT_CODE)
    cainiao_counts = Counter(df.loc[cainiao_mask, "派件网点简码"])
    sunyou_counts = Counter(df.loc[sunyou_mask, "派件网点简码"])
    aliexpress_count = int(cainiao_mask.sum())
    sunyou_count = int(sunyou_mask.sum())

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        template_file = find_latest_template_file()
        print(f"Using template file: {template_file}")
        wb = load_workbook(template_file, keep_links=False)

    write_data_source(wb, df)
    rebuild_non_auckland_pivot(wb, station_counts)
    rebuild_auckland_pivot(wb, route_counts)

    auckland_total = update_auckland_sheet(wb, route_counts)
    non_auckland_total = update_non_auckland_sheet(
        wb,
        station_counts,
        cainiao_counts,
        sunyou_counts,
        aliexpress_count,
        sunyou_count,
        auckland_total,
        len(df),
    )

    overlap = df[(df["派件网点简码"] != "AKL") & df["路由码"].isin(get_auckland_route_codes(wb))]

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_FILE
    try:
        wb.save(output_file)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_FILE.with_name(f"{OUTPUT_FILE.stem}_{timestamp}{OUTPUT_FILE.suffix}")
        wb.save(output_file)
        print("Default updated report is open or locked; wrote a new file instead.")

    print(f"Source rows: {len(df)}")
    print(f"Auckland route total: {auckland_total}")
    print(f"Non-Auckland station total: {non_auckland_total}")
    classified_total = auckland_total + non_auckland_total
    print(f"Workbook total logic: {len(df)}")
    print(f"Unassigned station total: {len(df) - classified_total}")
    print(f"Unique waybill total: {df['运单号'].nunique()}")
    merchant_counts = Counter(df[MERCHANT_COLUMN])
    print(
        "Merchant totals: "
        f"TEMU={merchant_counts.get(MERCHANT_CODES['TEMU'], 0)}, "
        f"Cainiao={merchant_counts.get(MERCHANT_CODES['CAINIAO'], 0)}, "
        f"Sunyou={merchant_counts.get(MERCHANT_CODES['SUNYOU'], 0)}"
    )
    if not overlap.empty:
        print("Potential double-count rows:")
        print(overlap[REQUIRED_COLUMNS].to_string(index=False))
    print(f"Done: {output_file}")


def get_auckland_route_codes(wb):
    ws = wb["奥克兰"]
    total_row = find_total_row(ws, column=1)
    return {
        clean_route_code(ws.cell(row, 1).value)
        for row in range(3, total_row)
        if clean_route_code(ws.cell(row, 1).value)
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update the daily report workbook.")
    parser.add_argument("--source-file", help="Use this exact center waybill query workbook")
    parser.add_argument(
        "--allow-old-source",
        action="store_true",
        help="Allow a center waybill query file whose modified date is not today",
    )
    args = parser.parse_args()
    main(args.source_file, allow_old_source=args.allow_old_source)
