from datetime import datetime
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

from report_config import BOARD_3L_CAPACITY, BOARD_5L_CAPACITY, SUPPLIER_ROUTE_GROUPS


REPORT_FILES = sorted(
    [
        path
        for path in Path(".").glob("*当日数据统计*.xlsx")
        if not path.name.startswith("~$")
    ],
    key=lambda path: path.stat().st_mtime,
)
UPDATED_REPORT_FILES = sorted(
    [
        path
        for path in Path("output").glob("当日数据统计*.xlsx")
        if not path.name.startswith("~$")
    ],
    key=lambda path: path.stat().st_mtime,
)
REPORT_FILE = REPORT_FILES[-1] if REPORT_FILES else (UPDATED_REPORT_FILES[-1] if UPDATED_REPORT_FILES else Path("当日数据统计.xlsx"))
OUTPUT_DIR = Path("output")
SUPPLIER_DIR = OUTPUT_DIR / "supplier"
PROVINCE_DIR = OUTPUT_DIR / "province"
SCALE = 2.5
today = datetime.now()
REPORT_DATE = f"{today.month}月{today.day:02d}日"

BLUE = "#8EAADB"
LIGHT_BLUE = "#BDD7EE"
YELLOW = "#FFF200"
TOTAL_BLUE = "#DDEBF7"
WHITE = "#FFFFFF"
BLACK = "#000000"
TEXT_ONLY_ROUTE_CODES = {"HST", "GSB"}
PROVINCE_STATIONS_BY_MESSAGE = {
    "WGR": ("WGR",),
    "HMT": ("HMT",),
    "PMN": ("PMN", "PMNV2", "PALMERSTON NORTHV2"),
    "RTR": ("RTR",),
    "TPO": ("TPO",),
    "TRG": ("TRG",),
    "NPL_HST": ("NPL", "HST"),
    "WLTV2": ("WLTV2",),
    "NPMV2": ("NPMV2", "NEW PLYMOUTHV2"),
    "WGU": ("WGU", "WHANGANUI"),
    "GSB": ("GSB", "GISBORNE"),
}


def load_font(size, bold=False):
    candidates = [
        Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf" if bold else "C:/Windows/Fonts/simsun.ttc"),
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return ImageFont.truetype(str(path), int(size * SCALE))
    return ImageFont.load_default()


FONT_TITLE = load_font(28, bold=True)
FONT_HEADER = load_font(20, bold=True)
FONT_BODY = load_font(18)
FONT_BODY_BOLD = load_font(18, bold=True)
FONT_TOTAL_NUMBER = load_font(24, bold=True)


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_route_code(value):
    return re.sub(r"\s+", "", clean_text(value))


def is_route_code(value):
    return value in TEXT_ONLY_ROUTE_CODES or any(char.isdigit() for char in value)


def clean_number(value):
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return 0
    if float(number).is_integer():
        return int(number)
    return round(float(number), 2)


def clean_board_value(value):
    text = clean_text(value)
    if "/" in text or "(" in text:
        return text
    return clean_number(value)


def parse_total_breakdown(value, province_count):
    text = clean_text(value)
    numbers = [clean_number(match) for match in re.findall(r"-?\d+(?:\.\d+)?", text)]
    if len(numbers) >= 4:
        return numbers[0], numbers[1], numbers[2], numbers[3]
    if len(numbers) >= 3:
        unassigned_count = max(0, numbers[0] - numbers[1] - numbers[2])
        return numbers[0], numbers[1], numbers[2], unassigned_count
    if len(numbers) == 1:
        return numbers[0], numbers[0] - province_count, province_count, 0
    return 0, 0, province_count, 0


def non_auckland_summary_rows(df, total_row):
    header_rows = df.index[df.iloc[:, 0].map(clean_text).eq("Aliexpress 单量")]
    header_row = int(header_rows[0]) if not header_rows.empty else total_row + 1
    return header_row, header_row + 1


def extract_board_forecast_table(df, title, default_capacity):
    header_rows = df.index[df.iloc[:, 8].map(clean_text).eq(title)]
    if header_rows.empty:
        raise ValueError(f"Could not find {title} in 非奥克兰")
    header_row = int(header_rows[0])

    total_rows = df.index[
        (df.index > header_row)
        & df.iloc[:, 7].map(clean_text).eq("总计")
    ]
    if total_rows.empty:
        raise ValueError(f"Could not find the total row for {title} in 非奥克兰")
    total_row = int(total_rows[0])

    table = df.iloc[header_row + 1 : total_row + 1, [7, 8]].copy()
    table.columns = ["station", "boards"]
    table["station"] = table["station"].map(clean_text)
    table["boards"] = table["boards"].map(clean_board_value)
    capacity = clean_number(df.iloc[header_row, 7]) or default_capacity
    return table, capacity


def text_size(draw, text, font):
    text = format_display(text)
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0], box[3] - box[1]


def px(value):
    return int(round(value * SCALE))


def scaled_box(values):
    return [px(value) for value in values]


def line_width():
    return max(1, int(round(SCALE)))


def format_display(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def draw_centered_text(draw, box, text, font, fill=BLACK):
    text = format_display(text)
    x1, y1, x2, y2 = scaled_box(box)
    width, height = text_size(draw, text, font)
    draw.text((x1 + (x2 - x1 - width) / 2, y1 + (y2 - y1 - height) / 2 - 1), text, font=font, fill=fill)


def draw_centered_fit_text(draw, box, text, font_size, bold=False, fill=BLACK, min_size=14, padding=10):
    text = format_display(text)
    x1, y1, x2, y2 = scaled_box(box)
    max_width = x2 - x1 - px(padding * 2)
    max_height = y2 - y1 - px(padding)
    size = font_size
    while size >= min_size:
        font = load_font(size, bold=bold)
        width, height = text_size(draw, text, font)
        if width <= max_width and height <= max_height:
            break
        size -= 1
    draw.text((x1 + (x2 - x1 - width) / 2, y1 + (y2 - y1 - height) / 2 - 1), text, font=font, fill=fill)


def draw_left_text(draw, box, text, font, fill=BLACK, padding=8):
    text = format_display(text)
    x1, y1, x2, y2 = scaled_box(box)
    _, height = text_size(draw, text, font)
    draw.text((x1 + px(padding), y1 + (y2 - y1 - height) / 2 - 1), text, font=font, fill=fill)


def rect(draw, coords, **kwargs):
    draw.rectangle(scaled_box(coords), **kwargs)


def render_table_image(
    path,
    title,
    headers,
    rows,
    widths,
    title_fill=BLUE,
    header_fill=LIGHT_BLUE,
    fit_body_columns=(),
    hide_total_label=False,
):
    row_h = 36
    title_h = 58
    header_h = 64
    margin = 0
    width = sum(widths) + margin * 2
    height = title_h + header_h + row_h * len(rows) + margin * 2

    image = Image.new("RGB", (px(width), px(height)), WHITE)
    draw = ImageDraw.Draw(image)

    y = margin
    rect(draw, [margin, y, width - margin - 1, y + title_h], fill=title_fill, outline=BLACK, width=line_width())
    draw_centered_fit_text(draw, (margin, y, width - margin, y + title_h), title, 28, bold=True)
    y += title_h

    x = margin
    for header, col_w in zip(headers, widths):
        rect(draw, [x, y, x + col_w, y + header_h], fill=header_fill, outline=BLACK, width=line_width())
        draw_centered_text(draw, (x, y, x + col_w, y + header_h), header, FONT_HEADER)
        x += col_w
    y += header_h

    for row in rows:
        is_total = clean_text(row[0]) in {"总计", "Total"}
        x = margin
        for idx, (value, col_w) in enumerate(zip(row, widths)):
            display_value = "" if is_total and hide_total_label and idx == 0 else value
            fill = TOTAL_BLUE if is_total else WHITE
            rect(draw, [x, y, x + col_w, y + row_h], fill=fill, outline=BLACK, width=line_width())
            font = FONT_BODY_BOLD if is_total else FONT_BODY
            if not is_total and idx in fit_body_columns:
                draw_centered_fit_text(
                    draw,
                    (x, y, x + col_w, y + row_h),
                    display_value,
                    18,
                    min_size=12,
                    padding=6,
                )
            else:
                draw_centered_text(
                    draw,
                    (x, y, x + col_w, y + row_h),
                    display_value,
                    font,
                )
            x += col_w
        y += row_h

    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path)


def supplier_display_rows(supplier, group):
    records = [
        {
            "route_code": clean_route_code(row.route_code),
            "quantity": clean_number(row.quantity),
        }
        for row in group.itertuples(index=False)
    ]
    route_positions = {}
    for index, record in enumerate(records):
        route_code = record["route_code"]
        if route_code in route_positions:
            raise ValueError(f"Duplicate route code for {supplier}: {route_code}")
        route_positions[route_code] = index

    configured_groups = []
    consumed_routes = set()
    for raw_members in SUPPLIER_ROUTE_GROUPS.get(supplier, []):
        members = tuple(clean_route_code(route_code) for route_code in raw_members)
        if len(members) < 2 or any(not member for member in members):
            raise ValueError(f"Invalid supplier route group for {supplier}: {raw_members}")
        if len(set(members)) != len(members):
            raise ValueError(f"Duplicate member in supplier route group for {supplier}: {members}")
        overlap = consumed_routes.intersection(members)
        if overlap:
            raise ValueError(
                f"Overlapping supplier route groups for {supplier}: {sorted(overlap)}"
            )
        missing = [member for member in members if member not in route_positions]
        if missing:
            raise ValueError(
                f"Missing configured routes for {supplier}: {', '.join(missing)}"
            )
        consumed_routes.update(members)
        configured_groups.append((min(route_positions[member] for member in members), members))

    group_by_first_position = {position: members for position, members in configured_groups}
    rows = []
    for index, record in enumerate(records):
        route_code = record["route_code"]
        members = group_by_first_position.get(index)
        if members:
            quantities = [records[route_positions[member]]["quantity"] for member in members]
            rows.append(
                [
                    " / ".join(members),
                    " / ".join(format_display(value) for value in quantities)
                    + f" ({format_display(sum(quantities))})",
                    supplier,
                ]
            )
        elif route_code not in consumed_routes:
            rows.append([route_code, record["quantity"], supplier])

    displayed_total = sum(record["quantity"] for record in records)
    return rows, displayed_total


def render_supplier_image(path, supplier, group):
    rows, displayed_total = supplier_display_rows(supplier, group)
    rows.append(["Total", int(group["quantity"].sum()), supplier])
    if displayed_total != group["quantity"].sum():
        raise ValueError(f"Supplier display total mismatch for {supplier}")
    has_grouped_routes = bool(SUPPLIER_ROUTE_GROUPS.get(supplier))
    render_table_image(
        path=path,
        title=f"{REPORT_DATE}奥克兰分单 - {supplier}",
        headers=["路由码\nroute code", "到件货量（预估）\nVolume of goods", "供应商\nsupplier"],
        rows=rows,
        widths=[300, 340, 260] if has_grouped_routes else [170, 260, 260],
        fit_body_columns=(0, 1) if has_grouped_routes else (),
        hide_total_label=False,
    )


def render_route_image(path, title, detail):
    rows = [[row.route_code, row.quantity] for row in detail.itertuples(index=False)]
    rows.append(["总计", int(detail["quantity"].sum())])
    render_table_image(
        path=path,
        title=title,
        headers=["Route Code", "Quantity"],
        rows=rows,
        widths=[300, 220],
        title_fill=YELLOW,
        header_fill=WHITE,
    )


def render_side_by_side_route_image(path, tables):
    title_h = 54
    row_h = 34
    gap = 120
    table_w = 520
    widths = [300, 220]
    max_rows = max(len(detail) + 1 for _, detail in tables)
    height = title_h + row_h * max_rows
    width = table_w * len(tables) + gap * (len(tables) - 1)

    image = Image.new("RGB", (px(width), px(height)), WHITE)
    draw = ImageDraw.Draw(image)

    for table_index, (title, detail) in enumerate(tables):
        x0 = table_index * (table_w + gap)
        rect(draw, [x0, 0, x0 + table_w, title_h], fill=YELLOW, outline=BLACK, width=line_width())
        draw_centered_fit_text(draw, (x0, 0, x0 + table_w, title_h), title, 20, bold=True)

        y = title_h
        rows = [[row.route_code, row.quantity] for row in detail.itertuples(index=False)]
        rows.append(["总计", int(detail["quantity"].sum())])

        for route_code, quantity in rows:
            is_total = route_code == "总计"
            fill = TOTAL_BLUE if is_total else WHITE
            rect(draw, [x0, y, x0 + widths[0], y + row_h], fill=fill if is_total else WHITE, outline=BLACK, width=line_width())
            rect(draw, [x0 + widths[0], y, x0 + table_w, y + row_h], fill=fill, outline=BLACK, width=line_width())
            font = FONT_BODY_BOLD if is_total else FONT_BODY
            draw_centered_text(draw, (x0, y, x0 + widths[0], y + row_h), route_code, font)
            draw_centered_text(draw, (x0 + widths[0], y, x0 + table_w, y + row_h), quantity, font)
            y += row_h

    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path)


def render_non_auckland_overview(
    path,
    overview,
    board_3l,
    board_5l,
    base_3l,
    base_5l,
    aliexpress_count,
    sunyou_count,
    total_count,
    auckland_count,
    province_count,
    unassigned_count,
):
    cell_h = 36
    title_h = 60
    header_h = 48
    gap = 28
    total_value_h = 52
    main_widths = [210, 210, 300, 300]
    side_widths = [100, 190]
    width = sum(main_widths) + gap + sum(side_widths)
    main_summary_y = title_h + header_h + cell_h * len(overview) + gap
    main_end_y = main_summary_y + header_h + max(cell_h, total_value_h)
    # Keep the board panel flush with the top edge; the main title remains in
    # the wider left panel while the board forecast starts on the first row.
    board_3l_y = 0
    board_3l_height = header_h + cell_h * len(board_3l)
    board_5l_y = board_3l_y + board_3l_height + gap
    board_5l_height = header_h + cell_h * len(board_5l)
    side_end_y = board_5l_y + board_5l_height
    height = max(main_end_y, side_end_y) + gap
    image = Image.new("RGB", (px(width), px(height)), WHITE)
    draw = ImageDraw.Draw(image)

    x0, y0 = 0, 0
    main_w = sum(main_widths)
    rect(draw, [x0, y0, x0 + main_w, y0 + title_h], fill=BLUE, outline=BLACK, width=line_width())
    draw_centered_text(draw, (x0, y0, x0 + main_w, y0 + title_h), f"{REPORT_DATE}非奥克兰到件货量", FONT_TITLE)
    y = y0 + title_h
    x = x0
    for header, col_w in zip(["派件网点", "到件货量", "外省菜鸟到货量", "外省顺友到货量"], main_widths):
        rect(draw, [x, y, x + col_w, y + header_h], fill=LIGHT_BLUE, outline=BLACK, width=line_width())
        draw_centered_text(draw, (x, y, x + col_w, y + header_h), header, FONT_HEADER)
        x += col_w
    y += header_h

    for row in overview.itertuples(index=False):
        values = [row.station, row.arrival_volume, row.cainiao_volume, row.sunyou_volume]
        is_total = row.station == "总计"
        x = x0
        for value, col_w in zip(values, main_widths):
            rect(draw, [x, y, x + col_w, y + cell_h], fill=TOTAL_BLUE if is_total else WHITE, outline=BLACK, width=line_width())
            draw_centered_text(draw, (x, y, x + col_w, y + cell_h), value, FONT_BODY_BOLD if is_total else FONT_BODY)
            x += col_w
        y += cell_h

    y += 28
    rect(draw, [x0, y, x0 + main_widths[0], y + header_h], fill=LIGHT_BLUE, outline=BLACK, width=line_width())
    draw_centered_text(draw, (x0, y, x0 + main_widths[0], y + header_h), "Aliexpress 单量", FONT_HEADER)
    rect(draw, [x0, y + header_h, x0 + main_widths[0], y + header_h + cell_h], fill=WHITE, outline=BLACK, width=line_width())
    draw_centered_text(draw, (x0, y + header_h, x0 + main_widths[0], y + header_h + cell_h), aliexpress_count, FONT_BODY)

    sunyou_x = x0 + main_widths[0]
    rect(draw, [sunyou_x, y, sunyou_x + main_widths[1], y + header_h], fill=LIGHT_BLUE, outline=BLACK, width=line_width())
    draw_centered_text(draw, (sunyou_x, y, sunyou_x + main_widths[1], y + header_h), "顺友单量", FONT_HEADER)
    rect(draw, [sunyou_x, y + header_h, sunyou_x + main_widths[1], y + header_h + cell_h], fill=WHITE, outline=BLACK, width=line_width())
    draw_centered_text(draw, (sunyou_x, y + header_h, sunyou_x + main_widths[1], y + header_h + cell_h), sunyou_count, FONT_BODY)

    total_x = x0 + main_widths[0] + main_widths[1]
    total_width = main_widths[2] + main_widths[3]
    if unassigned_count:
        total_display = (
            f"{total_count}（{auckland_count} + {province_count}"
            f" + {unassigned_count}未分配）"
        )
        total_header = "当天总量（奥克兰 + 外省 + 未分配）"
    else:
        total_display = f"{total_count}（{auckland_count} + {province_count}）"
        total_header = "当天总量（奥克兰 + 外省）"
    rect(draw, [total_x, y, total_x + total_width, y + header_h], fill="#FFC000", outline=BLACK, width=line_width())
    draw_centered_text(draw, (total_x, y, total_x + total_width, y + header_h), total_header, FONT_HEADER)
    rect(draw, [total_x, y + header_h, total_x + total_width, y + header_h + total_value_h], fill="#FFC000", outline=BLACK, width=line_width())
    draw_centered_text(draw, (total_x, y + header_h, total_x + total_width, y + header_h + total_value_h), total_display, FONT_TOTAL_NUMBER)

    def draw_board_table(x, y, title, data):
        rect(draw, [x, y, x + side_widths[0], y + header_h], fill=WHITE, outline=BLACK, width=line_width())
        draw_centered_text(draw, (x, y, x + side_widths[0], y + header_h), data.iloc[0]["base"], FONT_HEADER)
        rect(draw, [x + side_widths[0], y, x + sum(side_widths), y + header_h], fill=WHITE, outline=BLACK, width=line_width())
        draw_centered_text(draw, (x + side_widths[0], y, x + sum(side_widths), y + header_h), title, FONT_HEADER)
        y += header_h
        for row in data.itertuples(index=False):
            rect(draw, [x, y, x + side_widths[0], y + cell_h], fill=WHITE, outline=BLACK, width=line_width())
            draw_left_text(draw, (x, y, x + side_widths[0], y + cell_h), row.station, FONT_BODY, padding=6)
            rect(draw, [x + side_widths[0], y, x + sum(side_widths), y + cell_h], fill=WHITE, outline=BLACK, width=line_width())
            draw_centered_text(draw, (x + side_widths[0], y, x + sum(side_widths), y + cell_h), row.boards, FONT_BODY_BOLD if row.station == "总计" else FONT_BODY)
            y += cell_h

    side_x = main_w + gap
    board_3l_img = board_3l.copy()
    board_3l_img["base"] = base_3l
    draw_board_table(side_x, board_3l_y, "3L预测板数", board_3l_img)

    board_5l_img = board_5l.copy()
    board_5l_img["base"] = base_5l
    draw_board_table(side_x, board_5l_y, "5L预测板数", board_5l_img)

    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path)


def build_supplier_messages():
    df = pd.read_excel(REPORT_FILE, sheet_name="奥克兰", header=None, dtype=str).fillna("")
    rows = df.iloc[2:101, [0, 2, 6]].copy()
    rows.columns = ["route_code", "quantity", "supplier"]
    rows["route_code"] = rows["route_code"].map(clean_route_code)
    rows["supplier"] = rows["supplier"].map(clean_text)
    rows["quantity"] = rows["quantity"].map(clean_number)
    rows = rows[(rows["route_code"] != "") & (rows["supplier"] != "")]

    summary = []

    for supplier, group in rows.groupby("supplier", sort=True):
        total = int(group["quantity"].sum())

        safe_name = supplier.replace("/", "_").replace("\\", "_").replace(":", "_")

        render_supplier_image(SUPPLIER_DIR / f"{safe_name}.png", supplier, group)
        summary.append({"supplier": supplier, "routes": len(group), "total": total})

    return pd.DataFrame(summary)


def build_non_auckland_messages():
    df = pd.read_excel(REPORT_FILE, sheet_name="非奥克兰", header=None, dtype=str).fillna("")

    total_rows = df.index[df.iloc[:, 0].map(clean_text).eq("总计")]
    total_row = int(total_rows[0]) if not total_rows.empty else 10
    summary_header_row, summary_value_row = non_auckland_summary_rows(df, total_row)

    overview = df.iloc[2 : total_row + 1, [0, 2, 5, 6]].copy()
    overview.columns = ["station", "arrival_volume", "cainiao_volume", "sunyou_volume"]
    overview["station"] = overview["station"].map(clean_text)
    overview["arrival_volume"] = overview["arrival_volume"].map(clean_number)
    overview["cainiao_volume"] = overview["cainiao_volume"].map(clean_number)
    overview["sunyou_volume"] = overview["sunyou_volume"].map(clean_number)

    board_3l, base_3l = extract_board_forecast_table(
        df,
        "3L预测板数",
        BOARD_3L_CAPACITY,
    )
    board_5l, base_5l = extract_board_forecast_table(
        df,
        "5L预测板数",
        BOARD_5L_CAPACITY,
    )

    aliexpress_count = clean_number(df.iloc[summary_value_row, 0])
    sunyou_count = clean_number(df.iloc[summary_value_row, 2])
    province_total = overview.loc[overview["station"].eq("总计"), "arrival_volume"]
    province_count = clean_number(province_total.iloc[0]) if not province_total.empty else 0
    total_count, auckland_count, province_count, unassigned_count = (
        parse_total_breakdown(df.iloc[summary_value_row, 5], province_count)
    )
    if not total_count:
        total_count = clean_number(df.iloc[summary_value_row, 3]) or clean_number(
            df.iloc[summary_value_row, 5]
        )
        auckland_count = clean_number(df.iloc[summary_value_row, 4]) or total_count - province_count
        unassigned_count = max(
            0,
            total_count - auckland_count - province_count,
        )

    render_non_auckland_overview(
        PROVINCE_DIR / "非奥克兰总览.png",
        overview,
        board_3l,
        board_5l,
        base_3l,
        base_5l,
        aliexpress_count,
        sunyou_count,
        total_count,
        auckland_count,
        province_count,
        unassigned_count,
    )


def build_route_detail_messages():
    wb = load_workbook(REPORT_FILE, data_only=False, read_only=True)
    source_ws = wb["数据源1-预测"]
    route_counts_by_station = {}
    for route_code, station in source_ws.iter_rows(
        min_row=2, min_col=2, max_col=3, values_only=True
    ):
        route_code = clean_route_code(route_code)
        station = clean_text(station).upper()
        if route_code and station:
            station_counts = route_counts_by_station.setdefault(station, {})
            station_counts[route_code] = station_counts.get(route_code, 0) + 1

    details_by_name = {}

    for name, stations in PROVINCE_STATIONS_BY_MESSAGE.items():
        combined_counts = {}
        for station in stations:
            for route_code, quantity in route_counts_by_station.get(station, {}).items():
                combined_counts[route_code] = combined_counts.get(route_code, 0) + quantity
        rows = [
            {"route_code": route_code, "quantity": quantity}
            for route_code, quantity in sorted(combined_counts.items())
            if is_route_code(route_code)
        ]

        detail = pd.DataFrame(rows, columns=["route_code", "quantity"])
        details_by_name[name] = detail

        render_route_image(PROVINCE_DIR / f"{name}各线路预测.png", f"{REPORT_DATE}{name}各线路预测", detail)

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    supplier_summary = build_supplier_messages()
    build_non_auckland_messages()
    build_route_detail_messages()

    print(f"Generated supplier messages: {len(supplier_summary)}")
    print(f"Done: {OUTPUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
